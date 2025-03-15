from dataclasses import dataclass
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

@dataclass
class ExcelGeneratorConfig:
    """Excel生成器配置项"""
    str1_first_col_width: float = 14.3
    str1_default_col_width: float = 21.0
    str2_col_width: float = 30.0
    row_height: float = 35.0
    row_height_start: int = 3
    freeze_panes_row: int = 4
    merge_start_row: int = 1
    header_font: Font = Font(name='微软雅黑', bold=True, size=10)
    data_font: Font = Font(name='微软雅黑', size=9)
    color_a: PatternFill = PatternFill(start_color='FCE4D6', fill_type='solid')
    color_b: PatternFill = PatternFill(start_color='E2EFDA', fill_type='solid')
    str2_header_fill: PatternFill = PatternFill(start_color='DDEBF7', fill_type='solid')
    border_style: Border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

class ExcelGenerator:
    """Excel文件生成工具类"""
    
    def __init__(self, config: ExcelGeneratorConfig = None):
        self.config = config or ExcelGeneratorConfig()
        self.wb = None
        self.ws = None
        self._current_color = self.config.color_a
        self._previous_category = None

    def __enter__(self):
        """上下文管理器入口"""
        self.wb = Workbook()
        self.ws = self.wb.active
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        if self.wb:
            self.wb.close()

    def validate_input(self, json_data: Dict[str, Any]):
        """验证输入数据格式"""
        required_keys = {'baseinfo', 'str1', 'str2'}
        if not required_keys.issubset(json_data.keys()):
            raise ValueError("输入数据缺少必要字段，需要包含: baseinfo, str1, str2")

        if 'str1Num' not in json_data['baseinfo']:
            raise ValueError("baseinfo中缺少str1Num字段")

    def setup_worksheet(self):
        """初始化工作表基本配置"""
        # 设置冻结窗格
        self.ws.freeze_panes = f'A{self.config.freeze_panes_row}'
        # 设置默认对齐方式
        self.ws.sheet_properties.outlinePr.summaryBelow = False

    def _apply_cell_style(self, cell, is_header: bool = False):
        """应用单元格样式"""
        cell.font = self.config.header_font if is_header else self.config.data_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = self.config.border_style

    def process_str1_data(self, str1_data: Dict[str, list]):
        """处理STR1类型数据"""
        for row_idx, (h_key, values) in enumerate(sorted(str1_data.items(), key=lambda x: int(x[0][1:])), 1):
            # 写入数据并应用样式
            for col_idx, value in enumerate(values, 1):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_cell_style(cell, is_header=True)

            # 处理颜色交替逻辑
            self._handle_color_alternate(values[0], row_idx)

    def _handle_color_alternate(self, current_category: str, row_idx: int):
        """处理颜色交替逻辑"""
        if self._previous_category and current_category != self._previous_category:
            self._current_color = self.config.color_b if self._current_color == self.config.color_a else self.config.color_a
        self._previous_category = current_category

        # 应用整行颜色
        for col in range(1, 4):
            self.ws.cell(row=row_idx, column=col).fill = self._current_color

    def process_str2_data(self, str2_data: Dict[str, dict], str1_num: int):
        """处理STR2类型数据"""
        start_col = str1_num + 1
        for col_offset, (l_key, data) in enumerate(sorted(str2_data.items(), key=lambda x: int(x[0][1:])), start_col):
            self._process_str2_column(col_offset, data)

    def _process_str2_column(self, col: int, data: dict):
        """处理单个STR2列"""
        col_letter = get_column_letter(col)
        
        # 设置表头样式
        for row_idx in range(1, 6):
            cell = self.ws.cell(row=row_idx, column=col)
            cell.fill = self.config.str2_header_fill
            self._apply_cell_style(cell, is_header=True)

        # 写入数据
        for h_key, value in data.items():
            row_idx = int(h_key[1:])
            cell = self.ws.cell(row=row_idx, column=col, value=value)
            if row_idx > 5:
                self._apply_cell_style(cell)

    def configure_column_widths(self, str1_num: int, str2_count: int):
        """配置列宽"""
        for col in range(1, str1_num + str2_count + 1):
            width = self.config.str1_first_col_width if col == 1 else \
                    self.config.str1_default_col_width if col <= str1_num else \
                    self.config.str2_col_width
            self.ws.column_dimensions[get_column_letter(col)].width = width

    def save(self, output_file: str):
        """保存Excel文件"""
        try:
            self.wb.save(output_file)
        except PermissionError as e:
            raise IOError(f"文件{output_file}被其他程序占用，请关闭Excel窗口后重试") from e

    @classmethod
    def merge_duplicate_cells(cls, worksheet, start_row=1, end_row=None, 
                             start_col=1, end_col=None, align_merged=True):
        """合并重复单元格"""
        # 获取实际的结束行和列
        end_row = end_row or worksheet.max_row
        end_col = end_col or worksheet.max_column

        # 遍历每一列
        for col in range(start_col, end_col + 1):
            current_value = None
            merge_start = None

            # 遍历当前列的每一行
            for row in range(start_row, end_row + 1):
                cell = worksheet.cell(row=row, column=col)
                value = cell.value

                # 初始化第一个单元格
                if merge_start is None:
                    current_value = value
                    merge_start = row
                    continue

                # 检测到值变化时处理合并
                if value != current_value:
                    if row - 1 >= merge_start + 1:  # 需要合并的行数>=2
                        worksheet.merge_cells(
                            start_row=merge_start,
                            end_row=row - 1,
                            start_column=col,
                            end_column=col
                        )
                        if align_merged:
                            merged_cell = worksheet.cell(row=merge_start, column=col)
                            merged_cell.alignment = Alignment(vertical='center')
                    current_value = value
                    merge_start = row

            # 处理最后一批相同值
            if merge_start and end_row >= merge_start + 1:
                worksheet.merge_cells(
                    start_row=merge_start,
                    end_row=end_row,
                    start_column=col,
                    end_column=col
                )
                if align_merged:
                    merged_cell = worksheet.cell(row=merge_start, column=col)
                    merged_cell.alignment = Alignment(vertical='center')
    
def generate_excel(json_data: Dict[str, Any], output_file: str, config: ExcelGeneratorConfig = None):
    """快捷生成函数（保持兼容）"""
    try:
        with ExcelGenerator(config) as generator:
            # 数据验证
            generator.validate_input(json_data)
            
            # 初始化工作表
            generator.setup_worksheet()
            
            # 处理数据
            str1_num = int(json_data['baseinfo']['str1Num'])
            generator.process_str1_data(json_data['str1'])
            generator.process_str2_data(json_data['str2'], str1_num)
            
            # 配置表格
            generator.configure_column_widths(
                str1_num=str1_num,
                str2_count=len(json_data['str2'])
            )
            
            # 合并单元格
            ExcelGenerator.merge_duplicate_cells(
                worksheet=generator.ws,
                start_row=generator.config.merge_start_row,
                end_row=len(json_data['str1']),
                end_col=str1_num
            )
            
            # 保存文件
            generator.save(output_file)
    except Exception as e:
        print(f"生成Excel文件时发生错误: {str(e)}")
        raise  # 向上抛出异常由调用方处理


if __name__ == '__main__':
    # 从文件加载数据
    input_json = 'ai_note2ca_card/data/input/json/temp/flattened_output.json'
    output_xlsx = 'ai_note2ca_card/data/output/output5.xlsx'
    
    try:
        with open(input_json, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 使用默认配置生成Excel
        generate_excel(data, output_xlsx)
        
        print(f"Excel文件已成功生成至: {output_xlsx}")
    except FileNotFoundError:
        print(f"输入文件不存在: {input_json}")
        exit(1)
    except json.JSONDecodeError:
        print(f"JSON文件格式错误: {input_json}")
        exit(1)
